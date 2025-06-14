import datetime
import os
from typing import Any, List

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from database.orm import AsyncOrm
from schemas.categories_and_jobs import Subcategory, Category, Jobtype
from schemas.location import Location
from schemas.reports import OperationWithJobs
from schemas.search import TransportNumber
from schemas.users import User
from utils.date_time_service import convert_date_time
from utils.translator import translator as t
import pandas as pd


async def individual_mechanic_excel_report(operations: list[OperationWithJobs], mechanic_username: str,
                                           start_date: datetime.datetime, end_date: datetime.datetime,
                                           report_type: str, lang: str, session: Any) -> str:
    """Генерация excel отчета. Возвращает путь к файлу"""
    # Создаем директорию для отчетов, если ее нет
    os.makedirs("reports", exist_ok=True)

    # Цвета
    COLOR_DARK_BLUE = "203764"  # Темно-синий текст
    COLOR_LIGHT_BLUE = "D9E1F2"  # Светло-синий фон
    COLOR_LIGHT_GREEN = "E2EFDA"  # Светло-зеленый фон
    COLOR_LIGHT_YELLOW = "FFF2CC"  # Светло-желтый фон
    COLOR_LIGHT_RED = "FFCCCC"  # Светло-красный фон
    COLOR_LIGHT_GRAY = "F2F2F2"  # Светло-серый фон
    COLOR_WHITE = "FFFFFF"  # Белый фон

    # Стили границ
    BORDER_STYLE_THIN = "thin"
    BORDER_STYLE_MEDIUM = "medium"

    # Выравнивание
    ALIGN_CENTER = "center"
    ALIGN_LEFT = "left"
    ALIGN_RIGHT = "right"

    # Путь к файлу Excel в зависимости от типа отчета
    start_date = convert_date_time(start_date, with_tz=True)[0]
    end_date = convert_date_time(end_date, with_tz=True)[0]
    excel_path = f"reports/{report_type}_{start_date}_{end_date}.xlsx"
    sheet_name = "mechanic_report"
    title = f"📆 {await t.t(report_type, lang)} {start_date} - {end_date} {mechanic_username}"

    data = []
    max_len_a = 2
    max_len_b = 4
    max_len_c = 9
    max_len_d = 12
    max_len_e = 6
    max_len_f = 23
    max_len_g = 7
    max_len_h = 11

    columns = [
        "ID",
        await t.t('excel_date', lang),
        await t.t('excel_transport', lang),
        await t.t('excel_jobtype', lang),
        await t.t('excel_jobs', lang),
        await t.t('excel_avg_time', lang),
        await t.t('location', lang),
        await t.t('excel_comment', lang)
    ]
    data.append(columns)

    # формируем данные
    for operation in operations:
        for job in operation.jobs:
            location = await AsyncOrm.get_location_by_id(operation.location_id, session)
            row_data = [
                    # ID
                    f"{operation.id}",
                    # дата и время
                    f"{convert_date_time(operation.created_at, with_tz=True)[0]} {convert_date_time(operation.created_at, with_tz=True)[1]}",
                    # траспорт с категорией и номером
                    f"{await t.t(operation.transport_category, lang)} {operation.transport_subcategory}-{operation.transport_serial_number}",
                    # группа узлов
                    f"{await t.t(operation.jobs[0].jobtype_title, lang)}",
                    # работы
                    f"{await t.t(job.title, lang)}",
                    # среднее время на работу
                    f"{round(operation.duration / len(operation.jobs))}",
                    # локация
                    f"{await t.t(location.name, lang)}",
                    # комментарий
                    f"{operation.comment if operation.comment else '-'}"
                ]
            data.append(row_data)

            # записываем длину столбцов
            if len(row_data[0]) > max_len_a:
                max_len_a = len(row_data[0])

            if len(row_data[1]) > max_len_b:
                max_len_b = len(row_data[1])

            if len(row_data[2]) > max_len_c:
                max_len_c = len(row_data[2])

            if len(row_data[3]) > max_len_d:
                max_len_d = len(row_data[3])

            if len(row_data[4]) > max_len_e:
                max_len_e = len(row_data[4])

            if len(row_data[5]) > max_len_f:
                max_len_f = len(row_data[5])

            if len(row_data[6]) > max_len_g:
                max_len_g = len(row_data[6])

            if len(row_data[7]) > max_len_h:
                max_len_h = len(row_data[7])

    # Добавляем пустую строку между разделами
    data.append(["", "", "", "", "", "", ""])

    jobs_count = sum([len(operation.jobs) for operation in operations])
    duration_sum = str(sum([operation.duration for operation in operations]))
    # количество работ
    data.append([f"{await t.t('number_of_works', lang)}", "", "", f"{jobs_count}"])
    # общее среднее время
    data.append([f"{await t.t('excel_avg_time', lang)}:", "", "", f"{round(int(duration_sum)/jobs_count)} {await t.t('minutes', lang)}"])
    # общее время на работы
    data.append([f"{await t.t('total_time_spent', lang)}", "", "", f"{duration_sum} {await t.t('minutes', lang)}"])

    # Создаем DataFrame с заголовками
    df = pd.DataFrame(data)

    # Записываем в Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = max_len_a + 2
        worksheet.column_dimensions['B'].width = max_len_b
        worksheet.column_dimensions['C'].width = max_len_c
        worksheet.column_dimensions['D'].width = max_len_d
        worksheet.column_dimensions['E'].width = max_len_e + 2
        worksheet.column_dimensions['F'].width = max_len_f + 2
        worksheet.column_dimensions['G'].width = max_len_g + 2
        worksheet.column_dimensions['H'].width = 40

        # Заголовок отчета с улучшенным форматированием
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)  # Темно-синий текст

        # объединяем для заголовка
        worksheet.merge_cells('A1:H1')

        # делаем заголовок по центру
        title_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

        # Добавляем светло-синий фон для заголовка
        title_cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")

        # Добавляем границу для заголовка
        title_cell.border = Border(
            left=Side(style=BORDER_STYLE_MEDIUM),
            right=Side(style=BORDER_STYLE_MEDIUM),
            top=Side(style=BORDER_STYLE_MEDIUM),
            bottom=Side(style=BORDER_STYLE_MEDIUM)
        )
        for i in range(8):
            column_cell = worksheet.cell(row=1, column=i + 1)
            column_cell.border = Border(
                left=Side(style=BORDER_STYLE_MEDIUM),
                right=Side(style=BORDER_STYLE_MEDIUM),
                top=Side(style=BORDER_STYLE_MEDIUM),
                bottom=Side(style=BORDER_STYLE_MEDIUM)
            )

        # делаем названия колонок по центру и добавляем границы
        for i in range(8):
            column_cell = worksheet.cell(row=2, column=i+1)
            column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)
            column_cell.border = Border(
                left=Side(style=BORDER_STYLE_THIN),
                right=Side(style=BORDER_STYLE_THIN),
                bottom=Side(style=BORDER_STYLE_THIN)
            )

        # добавляем перенос текста для столбца комментарий
        start_row = 3
        end_row = jobs_count
        for i in range(start_row, end_row):
            column_cell = worksheet.cell(row=i, column=8)
            column_cell.alignment = Alignment(wrap_text=True)

        # объедениям столбцы у текста после таблицы
        row_number = jobs_count + 3
        for i in range(1, 4):
            worksheet.merge_cells(f'A{row_number + i}:C{row_number + i}')

    return excel_path


async def summary_mechanics_excel_report(start_date: datetime.datetime, end_date: datetime.datetime, report_type: str, lang: str,
                                         session: Any) -> str:
    """Генерация excel отчета. Возвращает путь к файлу"""
    # Создаем директорию для отчетов, если ее нет
    os.makedirs("reports", exist_ok=True)

    # Цвета
    COLOR_DARK_BLUE = "203764"  # Темно-синий текст
    COLOR_LIGHT_BLUE = "D9E1F2"  # Светло-синий фон
    COLOR_LIGHT_GREEN = "E2EFDA"  # Светло-зеленый фон
    COLOR_LIGHT_YELLOW = "FFF2CC"  # Светло-желтый фон
    COLOR_LIGHT_RED = "FFCCCC"  # Светло-красный фон
    COLOR_LIGHT_GRAY = "F2F2F2"  # Светло-серый фон
    COLOR_WHITE = "FFFFFF"  # Белый фон

    # Стили границ
    BORDER_STYLE_THIN = "thin"
    BORDER_STYLE_MEDIUM = "medium"

    # Выравнивание
    ALIGN_CENTER = "center"
    ALIGN_LEFT = "left"
    ALIGN_RIGHT = "right"

    # Путь к файлу Excel в зависимости от типа отчета
    start_date_str = convert_date_time(start_date, with_tz=True)[0]
    end_date_str = convert_date_time(end_date, with_tz=True)[0]
    excel_path = f"reports/{report_type}_{start_date_str}_{end_date_str}.xlsx"
    sheet_name = "summary_mechanics_report"
    title = f"📆 {await t.t(report_type, lang)} {start_date_str} - {end_date_str}"

    data = []

    columns = [
        await t.t('mechanic', lang),
        await t.t('works_count', lang),
        await t.t('works_time', lang),
        await t.t('avg_works', lang),
    ]

    data.append(columns)

    # формируем данные для отчета
    mechanics = await AsyncOrm.get_all_mechanics(session)

    works_count_rating = {}

    for idx, mechanic in enumerate(mechanics, start=1):
        row = list()

        row.append(mechanic.username)

        operations = await AsyncOrm.get_operations_for_user_by_period(mechanic.tg_id, start_date, end_date, session)

        # количество работ
        jobs_count = sum([len(operation.jobs) for operation in operations])
        row.append(f"{jobs_count}")

        # общее и среднее время
        duration_sum = sum([operation.duration for operation in operations])
        if jobs_count != 0:
            avg_time = round(int(duration_sum) / jobs_count)
        else:
            avg_time = 0
        row.append(f"{duration_sum}")
        row.append(f"{avg_time}")

        # записываем строчку в общий
        data.append(row)

        # запись для рейтинга
        works_count_rating[mechanic.username] = jobs_count

    # Добавляем пустую строку между разделами
    data.append(["", "", "", ""])

    # рейтинг механиков
    data.append([f"{await t.t('rating_works', lang)}", "", ""])

    sorted_mechanics = {k: v for k, v in sorted(works_count_rating.items(), key=lambda item: item[1], reverse=True)}
    for k, v in sorted_mechanics.items():
        data.append([f"{k}", f"{v}"])

    # Создаем DataFrame с заголовками
    df = pd.DataFrame(data)

    # Записываем в Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 25
        worksheet.column_dimensions['D'].width = 25

        # Заголовок отчета с улучшенным форматированием
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)  # Темно-синий текст

        # объединяем для заголовка
        worksheet.merge_cells('A1:D1')

        # делаем заголовок по центру
        title_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

        # Добавляем светло-синий фон для заголовка
        title_cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")

        # Добавляем границу для заголовка
        title_cell.border = Border(left=Side(style=BORDER_STYLE_MEDIUM), right=Side(style=BORDER_STYLE_MEDIUM),
            top=Side(style=BORDER_STYLE_MEDIUM), bottom=Side(style=BORDER_STYLE_MEDIUM))

        # делаем названия колонок по центру и добавляем границы
        for i in range(4):
            column_cell = worksheet.cell(row=2, column=i + 1)
            column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)
            column_cell.border = Border(left=Side(style=BORDER_STYLE_THIN), right=Side(style=BORDER_STYLE_THIN),
                bottom=Side(style=BORDER_STYLE_THIN))

    return excel_path


async def vehicle_report_by_transport_excel_report(
        operations: List[OperationWithJobs],
        start_date: datetime.datetime,
        end_date: datetime.datetime,
        report_type: str,
        report_subtype: str,
        lang: str,
        session: Any,
        transport: TransportNumber) -> str:

    """Отчет по номеру транспорта"""
    # Создаем директорию для отчетов, если ее нет
    os.makedirs("reports", exist_ok=True)

    # Цвета
    COLOR_DARK_BLUE = "203764"  # Темно-синий текст
    COLOR_LIGHT_BLUE = "D9E1F2"  # Светло-синий фон
    COLOR_LIGHT_GREEN = "E2EFDA"  # Светло-зеленый фон
    COLOR_LIGHT_YELLOW = "FFF2CC"  # Светло-желтый фон
    COLOR_LIGHT_RED = "FFCCCC"  # Светло-красный фон
    COLOR_LIGHT_GRAY = "F2F2F2"  # Светло-серый фон
    COLOR_WHITE = "FFFFFF"  # Белый фон

    # Стили границ
    BORDER_STYLE_THIN = "thin"
    BORDER_STYLE_MEDIUM = "medium"

    # Выравнивание
    ALIGN_CENTER = "center"
    ALIGN_LEFT = "left"
    ALIGN_RIGHT = "right"

    # Путь к файлу Excel в зависимости от типа отчета
    start_date = convert_date_time(start_date, with_tz=True)[0]
    end_date = convert_date_time(end_date, with_tz=True)[0]
    excel_path = f"reports/{report_type}_{start_date}_{end_date}.xlsx"
    sheet_name = "vehicle_report"

    title = f"📆 {await t.t(report_type, lang)} {start_date} - {end_date} {await t.t(report_subtype, lang)} {transport.subcategory_title}-{transport.serial_number}"

    data = []
    columns = [
        "ID",
        await t.t('excel_date', lang),
        await t.t('mechanic', lang),
        await t.t('location', lang),
        await t.t('works_time', lang),
        await t.t('excel_comment', lang),
        await t.t('excel_jobtype', lang),
        await t.t('excel_jobs', lang)
    ]
    data.append(columns)

    # формируем данные
    for operation in operations:
        mechanic = await AsyncOrm.get_user_by_tg_id(operation.tg_id, session)
        location = await AsyncOrm.get_location_by_id(operation.location_id, session)

        for job in operation.jobs:
            date, time = convert_date_time(operation.created_at, with_tz=True)
            data.append(
                [
                    # ID
                    f"{operation.id}",
                    # Дата и время
                    f"{date} {time}",
                    # Механик
                    f"{mechanic.username}",
                    # Локация
                    f"{await t.t(location.name, lang)}",
                    # Суммарное время работ
                    f"{operation.duration}",
                    # Комментарий
                    f"{operation.comment if operation.comment else '-'}",
                    # Группа узлов
                    f"{await t.t(job.jobtype_title, lang)}",
                    # работа
                    f"{await t.t(job.title, lang)}"
                ]
            )

    # Создаем DataFrame с заголовками
    df = pd.DataFrame(data)

    # Записываем в Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 10
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 25
        worksheet.column_dimensions['F'].width = 45
        worksheet.column_dimensions['G'].width = 35
        worksheet.column_dimensions['H'].width = 35

        # Заголовок отчета с улучшенным форматированием
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)  # Темно-синий текст

        # объединяем для заголовка
        worksheet.merge_cells('A1:H1')

        # делаем заголовок по центру
        title_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

        # Добавляем светло-синий фон для заголовка
        title_cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")

        # Добавляем границу для заголовка
        title_cell.border = Border(left=Side(style=BORDER_STYLE_MEDIUM), right=Side(style=BORDER_STYLE_MEDIUM),
            top=Side(style=BORDER_STYLE_MEDIUM), bottom=Side(style=BORDER_STYLE_MEDIUM))

        # делаем названия колонок по центру и добавляем границы
        for i in range(8):
            column_cell = worksheet.cell(row=2, column=i + 1)
            column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)
            column_cell.border = Border(left=Side(style=BORDER_STYLE_THIN), right=Side(style=BORDER_STYLE_THIN),
                bottom=Side(style=BORDER_STYLE_THIN))

    return excel_path


async def vehicle_report_by_subcategory_excel_report(
        operations: List[OperationWithJobs],
        start_date: datetime.datetime,
        end_date: datetime.datetime,
        report_type: str,
        report_subtype: str,
        lang: str,
        session: Any,
        subcategory: Subcategory) -> str:

    """Отчет по подкатегории транспорта"""
    # Создаем директорию для отчетов, если ее нет
    os.makedirs("reports", exist_ok=True)

    # Цвета
    COLOR_DARK_BLUE = "203764"  # Темно-синий текст
    COLOR_LIGHT_BLUE = "D9E1F2"  # Светло-синий фон
    COLOR_LIGHT_GREEN = "E2EFDA"  # Светло-зеленый фон
    COLOR_LIGHT_YELLOW = "FFF2CC"  # Светло-желтый фон
    COLOR_LIGHT_RED = "FFCCCC"  # Светло-красный фон
    COLOR_LIGHT_GRAY = "F2F2F2"  # Светло-серый фон
    COLOR_WHITE = "FFFFFF"  # Белый фон

    # Стили границ
    BORDER_STYLE_THIN = "thin"
    BORDER_STYLE_MEDIUM = "medium"

    # Выравнивание
    ALIGN_CENTER = "center"
    ALIGN_LEFT = "left"
    ALIGN_RIGHT = "right"

    # Путь к файлу Excel в зависимости от типа отчета
    start_date = convert_date_time(start_date, with_tz=True)[0]
    end_date = convert_date_time(end_date, with_tz=True)[0]
    excel_path = f"reports/{report_type}_{start_date}_{end_date}.xlsx"
    sheet_name = "vehicle_report"

    title = f"📆 {await t.t(report_type, lang)} {start_date} - {end_date} {await t.t(report_subtype, lang)} {subcategory.title}"

    data = []
    columns = [
        "ID",
        await t.t('excel_date', lang),
        await t.t('excel_transport', lang),
        await t.t('mechanic', lang),
        await t.t('location', lang),
        await t.t('works_time', lang),
        await t.t('excel_comment', lang),
        await t.t('excel_jobtype', lang),
        await t.t('excel_jobs', lang)
    ]
    data.append(columns)

    # формируем данные
    for operation in operations:
        mechanic = await AsyncOrm.get_user_by_tg_id(operation.tg_id, session)
        location = await AsyncOrm.get_location_by_id(operation.location_id, session)

        for job in operation.jobs:
            date, time = convert_date_time(operation.created_at, with_tz=True)
            data.append(
                [
                    # ID
                    f"{operation.id}",
                    # Дата и время
                    f"{date} {time}",
                    # Транспорт
                    f"{operation.transport_subcategory}-{operation.transport_serial_number}",
                    # Механик
                    f"{mechanic.username}",
                    # Локация
                    f"{await t.t(location.name, lang)}",
                    # Суммарное время работ
                    f"{operation.duration}",
                    # Комментарий
                    f"{operation.comment if operation.comment else '-'}",
                    # Группа узлов
                    f"{await t.t(job.jobtype_title, lang)}",
                    # работа
                    f"{await t.t(job.title, lang)}"
                ]
            )

    # Создаем DataFrame с заголовками
    df = pd.DataFrame(data)

    # Записываем в Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 10
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 30
        worksheet.column_dimensions['G'].width = 35
        worksheet.column_dimensions['H'].width = 35
        worksheet.column_dimensions['I'].width = 35

        # Заголовок отчета с улучшенным форматированием
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)  # Темно-синий текст

        # объединяем для заголовка
        worksheet.merge_cells('A1:I1')

        # делаем заголовок по центру
        title_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

        # Добавляем светло-синий фон для заголовка
        title_cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")

        # Добавляем границу для заголовка
        title_cell.border = Border(left=Side(style=BORDER_STYLE_MEDIUM), right=Side(style=BORDER_STYLE_MEDIUM),
            top=Side(style=BORDER_STYLE_MEDIUM), bottom=Side(style=BORDER_STYLE_MEDIUM))

        # делаем названия колонок по центру и добавляем границы
        for i in range(9):
            column_cell = worksheet.cell(row=2, column=i + 1)
            column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)
            column_cell.border = Border(left=Side(style=BORDER_STYLE_THIN), right=Side(style=BORDER_STYLE_THIN),
                bottom=Side(style=BORDER_STYLE_THIN))

    return excel_path


async def vehicle_report_by_category_excel_report(
        operations: List[OperationWithJobs],
        start_date: datetime.datetime,
        end_date: datetime.datetime,
        report_type: str,
        report_subtype: str,
        lang: str,
        session: Any,
        category_title: str) -> str:

    """Отчет по категории транспорта"""
    # Создаем директорию для отчетов, если ее нет
    os.makedirs("reports", exist_ok=True)

    # Цвета
    COLOR_DARK_BLUE = "203764"  # Темно-синий текст
    COLOR_LIGHT_BLUE = "D9E1F2"  # Светло-синий фон
    COLOR_LIGHT_GREEN = "E2EFDA"  # Светло-зеленый фон
    COLOR_LIGHT_YELLOW = "FFF2CC"  # Светло-желтый фон
    COLOR_LIGHT_RED = "FFCCCC"  # Светло-красный фон
    COLOR_LIGHT_GRAY = "F2F2F2"  # Светло-серый фон
    COLOR_WHITE = "FFFFFF"  # Белый фон

    # Стили границ
    BORDER_STYLE_THIN = "thin"
    BORDER_STYLE_MEDIUM = "medium"

    # Выравнивание
    ALIGN_CENTER = "center"
    ALIGN_LEFT = "left"
    ALIGN_RIGHT = "right"

    # Путь к файлу Excel в зависимости от типа отчета
    start_date = convert_date_time(start_date, with_tz=True)[0]
    end_date = convert_date_time(end_date, with_tz=True)[0]
    excel_path = f"reports/{report_type}_{start_date}_{end_date}.xlsx"
    sheet_name = "vehicle_report"

    title = f"📆 {await t.t(report_type, lang)} {start_date} - {end_date} {await t.t(report_subtype, lang)} {await t.t(category_title, lang)}"

    data = []
    columns = [
        "ID",
        await t.t('excel_date', lang),
        await t.t('excel_transport', lang),
        await t.t('mechanic', lang),
        await t.t('location', lang),
        await t.t('works_time', lang),
        await t.t('excel_comment', lang),
        await t.t('excel_jobtype', lang),
        await t.t('excel_jobs', lang)
    ]
    data.append(columns)

    # формируем данные
    for operation in operations:
        mechanic = await AsyncOrm.get_user_by_tg_id(operation.tg_id, session)
        location = await AsyncOrm.get_location_by_id(operation.location_id, session)

        for job in operation.jobs:
            date, time = convert_date_time(operation.created_at, with_tz=True)
            data.append(
                [
                    # ID
                    f"{operation.id}",
                    # Дата и время
                    f"{date} {time}",
                    # Транспорт
                    f"{operation.transport_subcategory}-{operation.transport_serial_number}",
                    # Механик
                    f"{mechanic.username}",
                    # Локация
                    f"{await t.t(location.name, lang)}",
                    # Суммарное время работ
                    f"{operation.duration}",
                    # Комментарий
                    f"{operation.comment if operation.comment else '-'}",
                    # Группа узлов
                    f"{await t.t(job.jobtype_title, lang)}",
                    # работа
                    f"{await t.t(job.title, lang)}"
                ]
            )

    # Создаем DataFrame с заголовками
    df = pd.DataFrame(data)

    # Записываем в Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 10
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 30
        worksheet.column_dimensions['G'].width = 35
        worksheet.column_dimensions['H'].width = 35
        worksheet.column_dimensions['I'].width = 35

        # Заголовок отчета с улучшенным форматированием
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)  # Темно-синий текст

        # объединяем для заголовка
        worksheet.merge_cells('A1:I1')

        # делаем заголовок по центру
        title_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

        # Добавляем светло-синий фон для заголовка
        title_cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")

        # Добавляем границу для заголовка
        title_cell.border = Border(left=Side(style=BORDER_STYLE_MEDIUM), right=Side(style=BORDER_STYLE_MEDIUM),
            top=Side(style=BORDER_STYLE_MEDIUM), bottom=Side(style=BORDER_STYLE_MEDIUM))

        # делаем названия колонок по центру и добавляем границы
        for i in range(9):
            column_cell = worksheet.cell(row=2, column=i + 1)
            column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)
            column_cell.border = Border(left=Side(style=BORDER_STYLE_THIN), right=Side(style=BORDER_STYLE_THIN),
                bottom=Side(style=BORDER_STYLE_THIN))

    return excel_path


async def categories_work_excel_report(
        jobtypes: List[Jobtype],
        start_date: datetime.datetime,
        end_date: datetime.datetime,
        report_type: str,
        lang: str,
        session: Any) -> str:

    """Отчет по категориям работ"""
    # Создаем директорию для отчетов, если ее нет
    os.makedirs("reports", exist_ok=True)

    # Цвета
    COLOR_DARK_BLUE = "203764"  # Темно-синий текст
    COLOR_LIGHT_BLUE = "D9E1F2"  # Светло-синий фон
    COLOR_LIGHT_GREEN = "E2EFDA"  # Светло-зеленый фон
    COLOR_LIGHT_YELLOW = "FFF2CC"  # Светло-желтый фон
    COLOR_LIGHT_RED = "FFCCCC"  # Светло-красный фон
    COLOR_LIGHT_GRAY = "F2F2F2"  # Светло-серый фон
    COLOR_WHITE = "FFFFFF"  # Белый фон

    # Стили границ
    BORDER_STYLE_THIN = "thin"
    BORDER_STYLE_MEDIUM = "medium"

    # Выравнивание
    ALIGN_CENTER = "center"
    ALIGN_LEFT = "left"
    ALIGN_RIGHT = "right"

    # Путь к файлу Excel в зависимости от типа отчета
    converted_start_date = convert_date_time(start_date, with_tz=True)[0]
    converted_end_date = convert_date_time(end_date, with_tz=True)[0]
    excel_path = f"reports/{report_type}_{converted_start_date}_{converted_end_date}.xlsx"
    sheet_name = "work_categories_report"

    title = f"📆 {await t.t(report_type, lang)} {converted_start_date} - {converted_end_date}"

    data = []
    rows_to_merge = []
    rows_counter = 1
    headers_rows = []

    for idx, jt in enumerate(jobtypes, start=1):
        # заголовок jobtype
        emoji = jt.emoji + " " if jt.emoji else ""
        jobtype_header = f"{emoji}{await t.t(jt.title, lang)}\n"

        data.append([jobtype_header[:-1], ""])
        rows_counter += 1
        rows_to_merge.append(rows_counter)
        headers_rows.append(rows_counter)

        # количество выполненных операций по категории
        jobs = await AsyncOrm.get_jobs_by_jobtype_with_operation(jt.id, start_date, end_date, session)

        # если работ нет
        if len(jobs) == 0:
            text = await t.t("no_works", lang)
            data.append([text])
            rows_counter += 1
            continue

        jobs_count = {}
        transport_count = {}
        mechanic_count = {}
        for job in jobs:
            if jobs_count.get(job.job_title):
                jobs_count[job.job_title] += 1
            else:
                jobs_count[job.job_title] = 1

            transport = f"{job.subcategory_title}-{job.serial_number}"
            if transport_count.get(transport):
                transport_count[transport] += 1
            else:
                transport_count[transport] = 1

            if mechanic_count.get(job.mechanic_username):
                mechanic_count[job.mechanic_username] += 1
            else:
                mechanic_count[job.mechanic_username] = 1

        # сортировка работ по количеству
        sorted_jobs = {k: v for k, v in sorted(jobs_count.items(), key=lambda item: item[1], reverse=True)}
        sorted_transport = {k: v for k, v in sorted(transport_count.items(), key=lambda item: item[1], reverse=True)}
        sorted_mechanics = {k: v for k, v in sorted(mechanic_count.items(), key=lambda item: item[1], reverse=True)}

        for k, v in sorted_jobs.items():
            data.append([
                f"{await t.t(k, lang)}",
                f"{v}"
            ])
            rows_counter += 1

        # самые частый транспорт
        most_often_transport_header = await t.t('most_recent_transport', lang)
        data.append([most_often_transport_header[:-1], ""])
        rows_counter += 1
        rows_to_merge.append(rows_counter)

        counter = 0
        for k, v in sorted_transport.items():
            if counter == 3:
                break

            data.append([
                f"{k}",
                f"{v}"
            ])
            counter += 1
            rows_counter += 1

        # самые частые механики в категории
        most_often_mechanics_header = await t.t('most_recent_mechanics', lang)
        data.append([most_often_mechanics_header[:-1], ""])
        rows_counter += 1
        rows_to_merge.append(rows_counter)

        counter = 0
        for k, v in sorted_mechanics.items():
            if counter == 3:
                break

            data.append([
                f"{k}",
                f"{v}"
            ])
            counter += 1
            rows_counter += 1

        # строка после jobtype
        data.append(["", ""])
        rows_counter += 1

    # Создаем DataFrame с заголовками
    df = pd.DataFrame(data)

    # Записываем в Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 35

        # Заголовок отчета с улучшенным форматированием
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)  # Темно-синий текст

        # объединяем для заголовка
        worksheet.merge_cells('A1:B1')

        # делаем заголовок по центру
        title_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

        # Добавляем светло-синий фон для заголовка
        title_cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")

        # Добавляем границу для заголовка
        title_cell.border = Border(left=Side(style=BORDER_STYLE_MEDIUM), right=Side(style=BORDER_STYLE_MEDIUM),
                                   top=Side(style=BORDER_STYLE_MEDIUM), bottom=Side(style=BORDER_STYLE_MEDIUM))

        # стили подзаголовков
        for row in rows_to_merge:
            column_cell = worksheet.cell(row=row, column=1)

            # объединение ячеек
            worksheet.merge_cells(f'A{row}:B{row}')
            # выравнивание по центру
            column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)
            # жирный шрифт
            column_cell.font = Font(bold=True)

        # стили заголовков
        for row in headers_rows:
            column_cell_1 = worksheet.cell(row=row, column=1)
            column_cell_2 = worksheet.cell(row=row, column=2)

            # желтый цвет
            column_cell_1.fill = PatternFill(start_color=COLOR_LIGHT_YELLOW, end_color=COLOR_LIGHT_YELLOW,
                                           fill_type="solid")
            # граница тонкая у двух ячеек
            column_cell_1.border = Border(left=Side(style=BORDER_STYLE_THIN), right=Side(style=BORDER_STYLE_THIN),
                                        top=Side(style=BORDER_STYLE_THIN), bottom=Side(style=BORDER_STYLE_THIN))
            column_cell_2.border = Border(left=Side(style=BORDER_STYLE_THIN), right=Side(style=BORDER_STYLE_THIN),
                                        top=Side(style=BORDER_STYLE_THIN), bottom=Side(style=BORDER_STYLE_THIN))

    return excel_path


async def locations_excel_report(
        operations: List[OperationWithJobs],
        start_date: datetime.datetime,
        end_date: datetime.datetime,
        report_type: str,
        lang: str,
        location: Location) -> str:

    """Отчет по местоположению"""
    # Создаем директорию для отчетов, если ее нет
    os.makedirs("reports", exist_ok=True)

    # Цвета
    COLOR_DARK_BLUE = "203764"  # Темно-синий текст
    COLOR_LIGHT_BLUE = "D9E1F2"  # Светло-синий фон
    COLOR_LIGHT_GREEN = "E2EFDA"  # Светло-зеленый фон
    COLOR_LIGHT_YELLOW = "FFF2CC"  # Светло-желтый фон
    COLOR_LIGHT_RED = "FFCCCC"  # Светло-красный фон
    COLOR_LIGHT_GRAY = "F2F2F2"  # Светло-серый фон
    COLOR_WHITE = "FFFFFF"  # Белый фон

    # Стили границ
    BORDER_STYLE_THIN = "thin"
    BORDER_STYLE_MEDIUM = "medium"

    # Выравнивание
    ALIGN_CENTER = "center"
    ALIGN_LEFT = "left"
    ALIGN_RIGHT = "right"

    # Путь к файлу Excel в зависимости от типа отчета
    start_date = convert_date_time(start_date, with_tz=True)[0]
    end_date = convert_date_time(end_date, with_tz=True)[0]
    excel_path = f"reports/{report_type}_{start_date}_{end_date}.xlsx"
    sheet_name = "location_report"

    title = f"📆 {await t.t(report_type, lang)} {await t.t(location.name, lang)} {start_date} - {end_date}"

    data = []
    columns = [
        await t.t('excel_transport', lang),
        await t.t('excel_date', lang),
    ]
    data.append(columns)

    # строки которые потом надо будет объеденить как подзаголовки (сразу добавляем первые подзаголовки)
    rows_for_merge = [3, 4]
    # сразу пропускаем титул, шапку таблицы, первую категорию и подкатегорию
    rows_counter = 4

    # заголовок первой категории
    data.append([
        f"{await t.t(operations[0].transport_category, lang)}",
        ""
    ])

    # заголовок первой подкатегории
    data.append([
        f"{await t.t('subcategory', lang)} {operations[0].transport_subcategory}",
        ""
    ])

    # для учета смены категорий и подкатегорий
    unique_transport = {}
    current_category = operations[0].transport_category
    current_subcategory = operations[0].transport_subcategory
    category_counter = 0
    subcategory_counter = 0
    total_counter = 0

    for idx, operation in enumerate(operations, start=1):
        if unique_transport.get(f"{operation.transport_subcategory}-{operation.transport_serial_number}"):
            continue
        else:
            # чтобы не повторялся транспорт из операций
            key = f"{operation.transport_subcategory}-{operation.transport_serial_number}"
            unique_transport[key] = True

            # смена подкатегории для разбивки
            if operation.transport_subcategory != current_subcategory:
                # смена категории для разбивки
                if operation.transport_category != current_category:
                    # записываем итоги категории и подкатегории
                    data.append([
                        f'{await t.t("total", lang)} {current_subcategory}',
                        f'{subcategory_counter} {await t.t("items", lang)}'
                    ])

                    data.append([
                        f'{await t.t("total", lang)} {await t.t(current_category, lang)}',
                        f'{category_counter} {await t.t("items", lang)}'
                    ])

                    # записываем заголовки новых категории и подкатегории
                    data.append([
                        f"{await t.t(operation.transport_category, lang)}",
                        ""
                    ])

                    # добавляем запись о мерже строк
                    rows_counter += 3
                    rows_for_merge.append(rows_counter)

                    data.append([
                        f"{await t.t('subcategory', lang)} {operation.transport_subcategory}",
                        "",
                    ])

                    # добавляем запись о мерже строк
                    rows_counter += 1
                    rows_for_merge.append(rows_counter)

                    # меняем текущие подкатегорию и категорию
                    current_category = operation.transport_category
                    current_subcategory = operation.transport_subcategory
                    category_counter = 0
                    subcategory_counter = 0

                # без смены категории
                else:
                    # записываем итоги подкатегории
                    data.append([
                        f'{await t.t("total", lang)} {current_subcategory}',
                        f'{subcategory_counter} {await t.t("items", lang)}'
                    ])

                    # записываем заголовок новой подкатегории
                    data.append([
                        f"{await t.t('subcategory', lang)} {operation.transport_subcategory}",
                        ""
                    ])

                    # добавляем запись о мерже строк
                    rows_counter += 1
                    rows_for_merge.append(rows_counter)

                    # меняем текущие подкатегорию
                    current_subcategory = operation.transport_subcategory
                    subcategory_counter = 0

            date, time = convert_date_time(operation.created_at, with_tz=True)
            # записываем транспорт
            data.append([
                f"{operation.transport_subcategory}-{operation.transport_serial_number}",
                f"{date} {time}"
            ])
            category_counter += 1
            subcategory_counter += 1
            total_counter += 1

            # увеличиваем счетчик строк
            rows_counter += 1

    # запись последних строк
    data.append([
        f'{await t.t("total", lang)} {current_subcategory}',
        f'{subcategory_counter} {await t.t("items", lang)}'
    ])

    data.append([
        f'{await t.t("total", lang)} {await t.t(current_category, lang)}',
        f'{category_counter} {await t.t("items", lang)}'
    ])

    # запись финального подсчета на складе
    data.append([
        f'{await t.t("total_on_warehouse", lang)}',
        f'{total_counter} {await t.t("items", lang)}'
    ])

    # Создаем DataFrame с заголовками
    df = pd.DataFrame(data)

    # Записываем в Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 35

        # Заголовок отчета с улучшенным форматированием
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)  # Темно-синий текст

        # объединяем для заголовка
        worksheet.merge_cells('A1:B1')

        # объединение подзаголовков категорий и подкатегорий
        for row in rows_for_merge:
            worksheet.merge_cells(f'A{row}:B{row}')

        # делаем заголовок по центру
        title_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

        # Добавляем светло-синий фон для заголовка
        title_cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")

        # Добавляем границу для заголовка
        title_cell.border = Border(left=Side(style=BORDER_STYLE_MEDIUM), right=Side(style=BORDER_STYLE_MEDIUM),
                                   top=Side(style=BORDER_STYLE_MEDIUM), bottom=Side(style=BORDER_STYLE_MEDIUM))

        # делаем названия колонок по центру и добавляем границы
        for i in range(2):
            column_cell = worksheet.cell(row=2, column=i + 1)
            column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)
            column_cell.border = Border(left=Side(style=BORDER_STYLE_THIN), right=Side(style=BORDER_STYLE_THIN),
                                        bottom=Side(style=BORDER_STYLE_THIN))

        # выравниваем заголовки категорий и подкатегорий
        for n in rows_for_merge:
            column_cell = worksheet.cell(row=n, column=1)
            column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

    return excel_path


async def inefficiency_excel_report(
        operations: List[OperationWithJobs],
        start_date: datetime.datetime,
        end_date: datetime.datetime,
        report_type: str,
        lang: str,
        period: str) -> str:
    """Отчет по неэффективности"""
    # Создаем директорию для отчетов, если ее нет
    os.makedirs("reports", exist_ok=True)

    # Цвета
    COLOR_DARK_BLUE = "203764"  # Темно-синий текст
    COLOR_LIGHT_BLUE = "D9E1F2"  # Светло-синий фон
    COLOR_LIGHT_GREEN = "E2EFDA"  # Светло-зеленый фон
    COLOR_LIGHT_YELLOW = "FFF2CC"  # Светло-желтый фон
    COLOR_LIGHT_RED = "FFCCCC"  # Светло-красный фон
    COLOR_LIGHT_GRAY = "F2F2F2"  # Светло-серый фон
    COLOR_WHITE = "FFFFFF"  # Белый фон

    # Стили границ
    BORDER_STYLE_THIN = "thin"
    BORDER_STYLE_MEDIUM = "medium"

    # Выравнивание
    ALIGN_CENTER = "center"
    ALIGN_LEFT = "left"
    ALIGN_RIGHT = "right"

    # Путь к файлу Excel в зависимости от типа отчета
    converted_start_date = convert_date_time(start_date, with_tz=True)[0]
    converted_end_date = convert_date_time(end_date, with_tz=True)[0]
    excel_path = f"reports/{report_type}_{converted_start_date}_{converted_end_date}.xlsx"
    sheet_name = "inefficiency_report"

    title = f"📆 {await t.t(report_type, lang)} {converted_start_date} - {converted_end_date}"

    data = []
    rows_for_merge = [2]
    rows_counter = 2
    rows_headers = []

    # первый подзаголовок
    text = await t.t('repeatable_jobs', lang)

    data.append([
        f"{text[:-1]}",
        "",
        ""
    ])

    # колонки первой таблицы
    columns = [
        await t.t("excel_transport", lang),
        await t.t("excel_jobs", lang),
        await t.t("excel_count", lang)
    ]
    data.append(columns)
    rows_counter += 1
    rows_headers.append(rows_counter)

    transport_jobs_count = {}
    for o in operations:
        for job in o.jobs:
            key = f"{o.transport_subcategory}-{o.transport_serial_number} {await t.t(job.title, lang)}"

            if transport_jobs_count.get(key):
                transport_jobs_count[key] += 1
            else:
                transport_jobs_count[key] = 1

    # сортируем по количеству
    sorted_jobs = {k: v for k, v in sorted(transport_jobs_count.items(), key=lambda item: item[1], reverse=True)}

    # учитываем работы повторяющиеся только больше определенного количества раз за период
    if period == "today" or period == "yesterday":
        frequent_works = 2
    elif period == "week":
        frequent_works = 7
    elif period == "month":
        frequent_works = 15
    else:
        frequent_works = (end_date - start_date).days

    for k, v in sorted_jobs.items():
        if v >= frequent_works:
            transport = k.split(" ")[0]
            job_title = " ".join(k.split(" ")[1:])
            data.append([
                f"{transport}",
                f"{job_title}",
                f"{v}"
            ])
            rows_counter += 1

    # отделитель до следующего раздела
    data.append(["", "", ""])
    rows_counter += 1

    # второй подзаголовок
    text = await t.t('no_comments', lang)
    data.append([
        f"{text[:-1]}",
        "",
        ""
    ])
    rows_counter += 1
    rows_for_merge.append(rows_counter)

    # колонки
    columns = [
        "ID",
        await t.t('excel_date', lang),
        await t.t('excel_transport', lang),
    ]
    data.append(columns)
    rows_counter += 1
    rows_headers.append(rows_counter)

    # операции без комментариев
    for o in operations:
        if not o.comment:
            date, time = convert_date_time(o.created_at, with_tz=True)
            data.append([
                f"{o.id}",
                f"{date} {time}",
                f"{o.transport_subcategory}-{o.transport_serial_number}"
            ])

    # Создаем DataFrame с заголовками
    df = pd.DataFrame(data)

    # Записываем в Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 25

        # Заголовок отчета с улучшенным форматированием
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color=COLOR_DARK_BLUE)  # Темно-синий текст

        # объединяем для заголовка
        worksheet.merge_cells('A1:C1')

        # объединение подзаголовков
        for row in rows_for_merge:
            worksheet.merge_cells(f'A{row}:C{row}')

        # делаем заголовок по центру
        title_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

        # Добавляем светло-синий фон для заголовка
        title_cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")

        # Добавляем границу для заголовка
        title_cell.border = Border(left=Side(style=BORDER_STYLE_MEDIUM), right=Side(style=BORDER_STYLE_MEDIUM),
                                   top=Side(style=BORDER_STYLE_MEDIUM), bottom=Side(style=BORDER_STYLE_MEDIUM))

        # делаем названия колонок по центру
        for i in rows_headers:
            for c in range(3):
                column_cell = worksheet.cell(row=i, column=c+1)
                column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

        # выравниваем подзаголовки
        for n in rows_for_merge:
            column_cell = worksheet.cell(row=n, column=1)
            column_cell.alignment = Alignment(horizontal=ALIGN_CENTER)

    return excel_path